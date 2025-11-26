"""
Export Dialog Component - Export results to multiple formats
"""

import customtkinter as ctk
from tkinter import filedialog, messagebox
import os
import json
import csv
from datetime import datetime
from .theme_manager import ThemeManager


class ExportDialog(ctk.CTkToplevel):
    """Dialog for exporting analysis results"""
    
    def __init__(self, parent, candidates, stats):
        super().__init__(parent)
        
        self.candidates = candidates
        self.stats = stats
        
        # Window configuration
        self.title("Export Results")
        self.geometry("500x500")
        self.transient(parent)
        self.grab_set()
        
        # Center the window
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f'{width}x{height}+{x}+{y}')
        
        self._create_ui()
    
    def _create_ui(self):
        """Create dialog UI"""
        # Title
        title = ctk.CTkLabel(
            self,
            text="📊 Export Analysis Results",
            font=ThemeManager.get_font("heading")
        )
        title.pack(pady=(20, 10))
        
        # Subtitle
        subtitle = ctk.CTkLabel(
            self,
            text="Choose your export format:",
            font=ThemeManager.get_font("body"),
            text_color=("#2d3748", "#a0aec0")
        )
        subtitle.pack(pady=(0, 20))
        
        # Options container
        options_frame = ctk.CTkFrame(self, fg_color="transparent")
        options_frame.pack(pady=10, padx=30, fill="both", expand=True)
        
        # CSV button
        csv_btn = ctk.CTkButton(
            options_frame,
            text="📄 Export as CSV",
            command=self._export_csv,
            height=50,
            font=ThemeManager.get_font("body_bold"),
            **ThemeManager.get_button_style("success")
        )
        csv_btn.pack(pady=8, fill="x")
        
        # JSON button
        json_btn = ctk.CTkButton(
            options_frame,
            text="📋 Export as JSON",
            command=self._export_json,
            height=50,
            font=ThemeManager.get_font("body_bold"),
            corner_radius=10,
            fg_color=ThemeManager.COLORS["info"],
            hover_color=ThemeManager.COLORS["info_dark"],
            text_color="#ffffff"
        )
        json_btn.pack(pady=8, fill="x")
        
        # Excel button
        excel_btn = ctk.CTkButton(
            options_frame,
            text="📊 Export as Excel (XLSX)",
            command=self._export_excel,
            height=50,
            font=ThemeManager.get_font("body_bold"),
            corner_radius=10,
            fg_color=ThemeManager.COLORS["primary"],
            hover_color=ThemeManager.COLORS["primary_dark"],
            text_color="#ffffff"
        )
        excel_btn.pack(pady=8, fill="x")
        
        # Plain Text button
        txt_btn = ctk.CTkButton(
            options_frame,
            text="📝 Export as Plain Text",
            command=self._export_txt,
            height=50,
            font=ThemeManager.get_font("body_bold"),
            corner_radius=10,
            fg_color=("#4a5568", "#718096"),
            hover_color=("#2d3748", "#4a5568"),
            text_color="#ffffff"
        )
        txt_btn.pack(pady=8, fill="x")
        
        # Cancel button
        cancel_btn = ctk.CTkButton(
            self,
            text="Cancel",
            command=self.destroy,
            height=40,
            font=ThemeManager.get_font("body"),
            corner_radius=10,
            fg_color="transparent",
            border_width=2,
            border_color=("#a0aec0", "#718096"),
            text_color=("#a0aec0", "#718096")
        )
        cancel_btn.pack(pady=(5, 20), padx=30, fill="x")
    
    def _export_csv(self):
        """Export results to CSV"""
        if not self.candidates:
            messagebox.showwarning("No Data", "No analysis data available to export.")
            self.destroy()
            return
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            initialfile=f"cv_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        )
        
        if not filename:
            return
        
        try:
            with open(filename, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(["Rank", "CV File", "Fit Score", "Recommendation", "Skills", "Experience"])
                
                for i, candidate in enumerate(self.candidates, 1):
                    skills = ", ".join(candidate.get("skills", []))
                    experience = ", ".join(candidate.get("experience", []))
                    writer.writerow([
                        i,
                        os.path.basename(candidate.get("cv_file", "")),
                        candidate.get("fit_score", ""),
                        candidate.get("recommendation", ""),
                        skills,
                        experience
                    ])
            
            messagebox.showinfo("Success", f"Results exported successfully to:\n{filename}")
            self.destroy()
        except Exception as e:
            messagebox.showerror("Export Failed", f"Failed to export CSV:\n{str(e)}")
    
    def _export_json(self):
        """Export results to JSON"""
        if not self.candidates:
            messagebox.showwarning("No Data", "No analysis data available to export.")
            self.destroy()
            return
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            initialfile=f"cv_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        )
        
        if not filename:
            return
        
        try:
            export_data = {
                "analysis_date": datetime.now().isoformat(),
                "total_candidates": len(self.candidates),
                "statistics": self.stats,
                "candidates": self.candidates
            }
            
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, indent=2, ensure_ascii=False)
            
            messagebox.showinfo("Success", f"Results exported successfully to:\n{filename}")
            self.destroy()
        except Exception as e:
            messagebox.showerror("Export Failed", f"Failed to export JSON:\n{str(e)}")
    
    def _export_excel(self):
        """Export results to Excel (XLSX)"""
        if not self.candidates:
            messagebox.showwarning("No Data", "No analysis data available to export.")
            self.destroy()
            return
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"cv_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        
        if not filename:
            return
        
        try:
            # Try to use openpyxl for Excel export
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                
                wb = Workbook()
                ws = wb.active
                ws.title = "CV Analysis Results"
                
                # Header row with styling
                headers = ["Rank", "CV File", "Fit Score", "Recommendation", "Skills", "Experience"]
                ws.append(headers)
                
                # Style header
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                
                # Data rows
                for i, candidate in enumerate(self.candidates, 1):
                    skills = ", ".join(candidate.get("skills", []))
                    experience = ", ".join(candidate.get("experience", []))
                    ws.append([
                        i,
                        os.path.basename(candidate.get("cv_file", "")),
                        candidate.get("fit_score", ""),
                        candidate.get("recommendation", ""),
                        skills,
                        experience
                    ])
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                wb.save(filename)
                messagebox.showinfo("Success", f"Results exported successfully to:\n{filename}")
                self.destroy()
                
            except ImportError:
                # Fall back to CSV if openpyxl is not available
                messagebox.showwarning(
                    "Excel Export Not Available",
                    "The 'openpyxl' library is not installed.\n\n"
                    "Exporting as CSV instead.\n\n"
                    "To enable Excel export, install: pip install openpyxl"
                )
                # Change extension to .csv
                filename = filename.replace('.xlsx', '.csv')
                self._save_as_csv(filename)
                
        except Exception as e:
            messagebox.showerror("Export Failed", f"Failed to export Excel:\n{str(e)}")
    
    def _export_txt(self):
        """Export results to Plain Text"""
        if not self.candidates:
            messagebox.showwarning("No Data", "No analysis data available to export.")
            self.destroy()
            return
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
            initialfile=f"cv_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        )
        
        if not filename:
            return
        
        try:
            with open(filename, 'w', encoding='utf-8') as f:
                f.write("=" * 80 + "\n")
                f.write("CV ANALYSIS RESULTS\n")
                f.write("=" * 80 + "\n\n")
                f.write(f"Analysis Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Total Candidates: {len(self.candidates)}\n\n")
                
                # Statistics
                f.write("STATISTICS:\n")
                f.write("-" * 40 + "\n")
                f.write(f"  Shortlisted: {self.stats.get('shortlist', 0)}\n")
                f.write(f"  For Review: {self.stats.get('review', 0)}\n")
                f.write(f"  Other: {self.stats.get('reject', 0)}\n\n")
                
                # Candidates
                f.write("CANDIDATES:\n")
                f.write("=" * 80 + "\n\n")
                
                for i, candidate in enumerate(self.candidates, 1):
                    f.write(f"#{i} - {os.path.basename(candidate.get('cv_file', 'Unknown'))}\n")
                    f.write("-" * 80 + "\n")
                    f.write(f"Fit Score: {candidate.get('fit_score', 'N/A')}%\n")
                    f.write(f"Recommendation: {candidate.get('recommendation', 'N/A')}\n")
                    
                    skills = candidate.get("skills", [])
                    if skills:
                        f.write(f"Skills: {', '.join(skills)}\n")
                    
                    experience = candidate.get("experience", [])
                    if experience:
                        f.write(f"Experience: {', '.join(experience)}\n")
                    
                    f.write("\n")
            
            messagebox.showinfo("Success", f"Results exported successfully to:\n{filename}")
            self.destroy()
        except Exception as e:
            messagebox.showerror("Export Failed", f"Failed to export text file:\n{str(e)}")
    
    def _save_as_csv(self, filename):
        """Helper method to save as CSV (used by Excel fallback)"""
        try:
            with open(filename, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(["Rank", "CV File", "Fit Score", "Recommendation", "Skills", "Experience"])
                
                for i, candidate in enumerate(self.candidates, 1):
                    skills = ", ".join(candidate.get("skills", []))
                    experience = ", ".join(candidate.get("experience", []))
                    writer.writerow([
                        i,
                        os.path.basename(candidate.get("cv_file", "")),
                        candidate.get("fit_score", ""),
                        candidate.get("recommendation", ""),
                        skills,
                        experience
                    ])
            
            messagebox.showinfo("Success", f"Results exported as CSV to:\n{filename}")
            self.destroy()
        except Exception as e:
            messagebox.showerror("Export Failed", f"Failed to export CSV:\n{str(e)}")
